import openpyxl

class Input:
    def __init__(self, sheet):
        self.sheet = sheet

    def read(self):
        wb_obj = openpyxl.load_workbook(self.sheet)
        sheet_obj = wb_obj.active
        max_row = sheet_obj.max_row
        
        products = []
        for i in range(2, max_row+1):
            cell_obj = sheet_obj.cell(row = i, column = 1)
            products.append(cell_obj.value)

        return products

    def write(self, status):
        wb_obj = openpyxl.load_workbook(self.sheet)
        sheet_obj = wb_obj.active
        max_row = sheet_obj.max_row
        
        for i in range(2, max_row+1):
            cell_obj = sheet_obj.cell(row = i, column = 2)
            cell_obj.value = status[i-2]

# Give the location of the file
# path = "C:\\Users\\Admin\\Desktop\\demo.xlsx"

# # To open the workbook
# # workbook object is created
# wb_obj = openpyxl.load_workbook(path)

# # Get workbook active sheet object
# # from the active attribute
# sheet_obj = wb_obj.active

# # Cell objects also have a row, column,
# # and coordinate attributes that provide
# # location information for the cell.

# # Note: The first row or
# # column integer is 1, not 0.

# # Cell object is created by using
# # sheet object's cell() method.
# cell_obj = sheet_obj.cell(row = 1, column = 1)

# # Print value of cell object
# # using the value attribute
# print(cell_obj.value)
